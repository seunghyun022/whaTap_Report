import argparse
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import *


ampm=["오전","오후"]
time=["12","1","2","3","4","5","6","7","8","9","10","11"]
rest_of_time=":00:00"
col="BCDEFGHIJKLMNOPQRSTUVWXYZ"


def AutoFitColumnSize(worksheet, columns=None, margin=2):
    for i, column_cells in enumerate(worksheet.columns):
        is_ok = False
        if columns == None:
            is_ok = True
        elif isinstance(columns, list) and i in columns:
            is_ok = True

        if is_ok:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin

    return worksheet

def format_csv(buf,tgt):
    gl_idx = 1
    day_set = set()
    for i in range(1, len(buf)):
        day_set.add(buf[i][0].split()[0])
    day_list = list(day_set)
    day_list.sort()
    day_list.remove(day_list[7])
    print(len(buf[0]))
    for name in range(1,len(buf[0])):
        tgt.merge_cells('A{0}:F{1}'.format(gl_idx,gl_idx+1))
        tgt['A{0}'.format(gl_idx)].alignment = Alignment('center','center')
        tgt['A{0}'.format(gl_idx)].border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='double'))
        tgt['A{0}'.format(gl_idx)].font=Font(bold=True)
        tgt['A{0}'.format(gl_idx)].fill = PatternFill(patternType='solid',fgColor=Color('D9D9D9'))
        tgt['A{0}'.format(gl_idx)]=buf[0][name]

        for day_idx in range(0,len(day_list)):
            date_idx = gl_idx + 2 + 4*day_idx
            tgt.merge_cells('A{0}:A{1}'.format(date_idx, date_idx + 3))
            tgt['A{0}'.format(date_idx)].alignment=Alignment('center','center')
            tgt['A{0}'.format(date_idx)]=day_list[day_idx]
            for time_idx in range(len(time)):
                tgt['{0}{1}'.format(col[time_idx],date_idx)]=ampm[0]+" "+time[time_idx]+rest_of_time
                tgt['{0}{1}'.format(col[time_idx],date_idx)].alignment=Alignment('right','center')
                tgt['{0}{1}'.format(col[time_idx], date_idx + 2)] = ampm[1] + " " + time[time_idx] + rest_of_time
                tgt['{0}{1}'.format(col[time_idx], date_idx + 2)].alignment=Alignment('right','center')

                query1 = tgt['A{0}'.format(date_idx)].value + " " + tgt['{0}{1}'.format(col[time_idx],date_idx)].value
                query2 = tgt['A{0}'.format(date_idx)].value + " " + tgt['{0}{1}'.format(col[time_idx], date_idx+2)].value

                q1 = False
                q2 = False

                for suspect in buf:
                    if query1 == suspect[0]:
                        value = suspect[name]
                        if value == '':
                            value = '0'
                            print(date_idx)

                        tgt['{0}{1}'.format(col[time_idx], date_idx + 1)] = int(value)
                        tgt['{0}{1}'.format(col[time_idx], date_idx + 1)].alignment=Alignment('right','center')
                        q1 = True
                    elif query2 == suspect[0]:
                        value = suspect[name]
                        if value == '':
                            value = '0'
                            print(date_idx)

                        tgt['{0}{1}'.format(col[time_idx], date_idx + 3)] = int(value)
                        tgt['{0}{1}'.format(col[time_idx], date_idx + 3)].alignment=Alignment('right','center')
                        q2 = True
                    elif q1 and q2:
                        break
        gl_idx=gl_idx+30
def format_csv_avg(buf,tgt):
    gl_idx = 1
    for i in "ABCD":
        tgt['{0}{1}'.format(i,gl_idx)].alignment = Alignment('center', 'center')
        tgt['{0}{1}'.format(i,gl_idx)].border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                   top=Side(style='thin'), bottom=Side(style='double'))
        tgt['{0}{1}'.format(i,gl_idx)].font = Font(bold=True)
        tgt['{0}{1}'.format(i,gl_idx)].fill = PatternFill(patternType='solid', fgColor=Color('D9D9D9'))
    tgt['A1'] = "인스턴스명"
    tgt['B1'] = "주간 평균 동시접속자수"
    tgt['C1'] = "주간 최대 동시접속자수"
    tgt['D1'] = "주간 최저 동시접속자수"
    gl_idx=gl_idx+1

    day_set = set()
    print(len(buf[0]))
    for i in range(1, len(buf)):
        day_set.add(buf[i][0].split()[0])
    day_list = list(day_set)
    day_list.sort()
    for i in range(len(buf) - 1, -1,-1):
        if day_list[len(day_list) - 1] in buf[i][0]:
            buf.remove(buf[i])

    for name in range(1,len(buf[0])):
        tgt['A{0}'.format(gl_idx)]=buf[0][name].split()[0]
        MAX=0
        MIN=int(0x8FFFFFFF)
        nr_entry=0
        sum=0
        for i in range(1, len(buf)):
            if buf[i][name] !="":
                MAX=max(MAX,int(buf[i][name]))
                MIN=min(MIN,int(buf[i][name]))
                nr_entry=nr_entry+1
                sum=sum+int(buf[i][name])
        tgt['B{0}'.format(gl_idx)] = int(sum/nr_entry)
        tgt['C{0}'.format(gl_idx)] = MAX
        tgt['D{0}'.format(gl_idx)] = MIN
        gl_idx=gl_idx+1

def writecsv (buf,filename):
    buf.save(filename)

def open_csv(file_name):
    buf=list()
    with open(file_name,newline='') as csvfile:
        reader=csv.reader(csvfile)
        for row in reader:
            buf.append(row)
    csvfile.close()
    return buf

def getargs():
    parser = argparse.ArgumentParser()
    parser.add_argument('InputFileName', type=str, help="Input File Name to be fromatted")
    args=parser.parse_args()

    return args.InputFileName

def main(file):
    file_name=file
    #wb = Workbook()
    #ws = wb.active
    avgb = Workbook()
    avgs = avgb.active
    file_content = open_csv("./inputs/"+file_name)
    #format_csv(file_content,ws)
    #writecsv(wb,"./outputs/result_{0}".format(file_name.split('.')[0]+".xls"))
    format_csv_avg(file_content,avgs)
    AutoFitColumnSize(avgs)
    writecsv(avgb,"./outputs/avg_{0}".format(file_name.split('.')[0]+".xls"))
def sub_main():
    file_list=os.listdir(os.getcwd()+"/inputs")
    file_list.remove(".DS_Store")
    for i in file_list:
        print(i)
        main(i)
if __name__ =='__main__':
    sub_main()